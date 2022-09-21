import re,os,xml, json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl.styles import Alignment

class TranslatedString:
    def __init__(self, handle, obj):
        self.text = obj["ReferenceText"]
        self.description = obj["ContextDescription"] if "ContextDescription" in obj else ""
        self.handle = handle

class SheetColumn:
    def __init__(self, name, width, locked:bool=True, wrap_text:bool=True, font=None):
        self.name = name
        self.width = width
        self.locked = locked
        self.wrap_text = wrap_text
        self.font = font

class Sheet:
    ROW_HEIGHT = 30

    def __init__(self, sheet, name):
        self.name = name
        self.sheet = sheet
        self.column_defs = []
        self._initialized = False
        self.filled_rows = 0

        self.sheet.protection.sheet = True

    def addColumnDefinition(self, column:SheetColumn):
        self.column_defs.append(column)

        # Set width
        self.sheet.column_dimensions[get_column_letter(len(self.column_defs))].width = column.width

        # Protection is set when adding cell content instead.

    def append(self, values):
        self.sheet.append(values)

    def addRow(self, values:list):
        if not self._initialized:
            self.sheet.append([column.name for column in self.column_defs])

            self._initialized = True

        self.filled_rows = self.filled_rows + 1

        if len(values) != len(self.column_defs):
            raise ValueError("Iterable length mismatch with column defs amount")

        self.sheet.append(values)

        for i in range(len(values)):
            column_def = self.column_defs[i]
            letter = get_column_letter(i + 1)

            cell = self.sheet[letter + str(self.filled_rows + 1)]

            self.sheet.row_dimensions[self.filled_rows + 1].height = Sheet.ROW_HEIGHT
            
            # Set font
            cell.font = column_def.font

            # Set alignment
            if column_def.wrap_text:
                cell.alignment = Alignment(wrap_text=True)

            # Unlock cell for editing
            if not column_def.locked:
                cell.protection = Protection(locked=False)


class Workbook:
    def __init__(self):
        self.book = openpyxl.Workbook()
        self.sheets = []
        self._initalized = False

    def addSheet(self, name):
        sheet = Sheet(self.book.create_sheet(name), name)
        self.sheets.append(sheet)

        if not self._initalized:
            self.book.remove(self.book["Sheet"]) # TODO maybe don't?
            self._initalized = True

        return sheet

    def save(self, file_name):
        self.book.save(file_name)

class Module:
    COLUMN_DEFINITIONS = [
        SheetColumn("Handle", 10, wrap_text=False, font=Font(color="999999")),
        SheetColumn("Original Text", 65),
        SheetColumn("Contextual Info", 40),
        SheetColumn("Translation", 65, locked=False),
        SheetColumn("Translation Notes", 30, locked=False),
        SheetColumn("Translation Author", 30, locked=False),
    ]

    def __init__(self, name):
        self.name = name
        self.translated_strings = {}

    def addTSK(self, tsk):
        self.translated_strings[tsk.handle] = tsk

    def export(self, book:Workbook):
        sheet = book.addSheet(self.name)

        for column_def in Module.COLUMN_DEFINITIONS:
            sheet.addColumnDefinition(column_def)

        for handle,tsk in self.translated_strings.items():
            sheet.addRow([
                handle,
                tsk.text,
                tsk.description,
                "", # Translation
                "", # Translation author
                "", # Translation notes
            ])

def importLocalization(file_name):
    file = open(file_name, "r")
    localization = json.load(file)

    # TODO use one module per feature? maybe just a column for it is enough
    module = Module(localization["ModTable"])

    for handle,data in localization["TranslatedStrings"].items():
        tsk = TranslatedString(handle, data)

        module.addTSK(tsk)

    book = Workbook()
    module.export(book)
    
    print("Saving workbook")
    book.save("output.xlsx")

importLocalization(r"C:\Users\Usuario\Documents\Larian Studios\Divinity Original Sin 2 Definitive Edition\Osiris Data\Epip\localization_template.json")